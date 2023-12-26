from openpyxl import load_workbook
from apps.kinoprincipal.models import Kinodb, Rekinodb, Chanchitodb, Combodb, Chao1db, Chao2db, Chao3db, Archivosxl
import datetime


def Importar_datos(archivo):
    # Leer Excel
    wb = load_workbook(filename=archivo)
    ws = wb["Kino_Int"]

    ultima_fila = ws.max_row - 3
    #print(ultima_fila)
    # Parametros Kino
    rango_casillas_kino = "C808:P" + str(ultima_fila)
    int_inicio_kino = 808
    columna_ganadores_kino = "LD"

    # Parametros Rekino
    rango_casillas_rekino = "AD872:AQ" + str(ultima_fila)
    int_inicio_rekino = 872
    columna_ganadores_rekino = "LF"

    # Parametros Chanchito
    rango_casillas_chanchito = "AT1645:BG" + str(ultima_fila)
    int_inicio_chanchito = 1645
    columna_ganadores_chanchito = "LG"

    # Parametros Combo
    rango_casillas_combo = "BX1645:CK" + str(ultima_fila)
    int_inicio_combo = 1645
    columna_ganadores_combo = "LI"

    # Parametros Chao Jefe 1
    rango_casillas_chao1 = "JK2371:JX" + str(ultima_fila)
    int_inicio_chao1 = 2371
    columna_ganadores_chao1 = "LM"

    # Parametros Chao Jefe 2
    rango_casillas_chao2 = "JZ2371:KM" + str(ultima_fila)
    int_inicio_chao2 = 2371
    columna_ganadores_chao2 = "LN"

    # Parametros Chao Jefe 3
    rango_casillas_chao3 = "KO2774:LB" + str(ultima_fila)
    int_inicio_chao3 = 2774
    columna_ganadores_chao3 = "LO"

    excel_a_modelo(ws, Kinodb, rango_casillas_kino,
                   int_inicio_kino, columna_ganadores_kino)
    excel_a_modelo(ws, Rekinodb, rango_casillas_rekino,
                   int_inicio_rekino, columna_ganadores_rekino)
    excel_a_modelo(ws, Chanchitodb, rango_casillas_chanchito,
                   int_inicio_chanchito, columna_ganadores_chanchito)
    excel_a_modelo(ws, Combodb, rango_casillas_combo,
                   int_inicio_combo, columna_ganadores_combo)
    excel_a_modelo(ws, Chao1db, rango_casillas_chao1,
                   int_inicio_chao1, columna_ganadores_chao1)
    excel_a_modelo(ws, Chao2db, rango_casillas_chao2,
                   int_inicio_chao2, columna_ganadores_chao2)
    excel_a_modelo(ws, Chao3db, rango_casillas_chao3,
                   int_inicio_chao3, columna_ganadores_chao3)


def excel_a_modelo(ws, Modelo, rango_casillas, int_inicio, columna_ganadores):

    lista_total = []
    contador = int_inicio
    for row in ws[rango_casillas]:
        str_contador = str(contador)
        sorteo = ws["A" + str_contador].value
        fecha = ws["B" + str_contador].value

        numero1 = int(row[0].value)
        numero2 = int(row[1].value)
        numero3 = int(row[2].value)
        numero4 = int(row[3].value)
        numero5 = int(row[4].value)
        numero6 = int(row[5].value)
        numero7 = int(row[6].value)
        numero8 = int(row[7].value)
        numero9 = int(row[8].value)
        numero10 = int(row[9].value)
        numero11 = int(row[10].value)
        numero12 = int(row[11].value)
        numero13 = int(row[12].value)
        numero14 = int(row[13].value)
        ganador = int(ws[columna_ganadores + str_contador].value)
        tipo_ingreso = "Excel"

        elemento = Modelo(id_sorteo=sorteo, fecha=fecha, number1=numero1, number2=numero2, number3=numero3, number4=numero4, number5=numero5, number6=numero6,
                          number7=numero7, number8=numero8, number9=numero9, number10=numero10, number11=numero11, number12=numero12, number13=numero13, number14=numero14, num_ganadores=ganador, tipo_ingreso=tipo_ingreso)
        lista_total.append(elemento)
        contador = contador + 1

    Modelo.objects.bulk_create(lista_total)
